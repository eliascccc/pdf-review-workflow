from __future__ import annotations

# ==========================================
# Imports
# ==========================================

import io
import mimetypes
import os
import platform
import re
import shutil
import subprocess
import time
import uuid
from dataclasses import dataclass, field
from email import policy
from email.parser import BytesParser
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pdfminer.high_level import extract_text
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import Color
from reportlab.pdfgen import canvas
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

try:
    import extract_msg  # type: ignore
except ImportError:
    extract_msg = None

from config import SQLALCHEMY_DATABASE_URI
from flask_app import Job


# ==========================================
# Paths / folders
# ==========================================

BASE_DIR = Path(__file__).resolve().parent

UPLOADS_DIR = BASE_DIR / "uploads"
WORK_DIR = BASE_DIR / "work"
INBOX_DIR = WORK_DIR / "incoming"
MAIL_ATTACHMENTS_DIR = WORK_DIR / "mail_attachments"
REVIEW_DIR = WORK_DIR / "review"
OUTPUT_DIR = BASE_DIR / "static" / "client"

DEFAULT_EXCEL_TEMPLATE = BASE_DIR / "mall.xlsx"
DEFAULT_OUTPUT_SHEET = "Sheet1"

EXCEL_PATH_TXT = BASE_DIR / "excel_File_and_Path.txt"
SUMMARY_PATH_TXT = BASE_DIR / "summary_File_and_Path.txt"

for folder in [
    UPLOADS_DIR,
    WORK_DIR,
    INBOX_DIR,
    MAIL_ATTACHMENTS_DIR,
    REVIEW_DIR,
    OUTPUT_DIR,
]:
    folder.mkdir(parents=True, exist_ok=True)


# ==========================================
# Database
# ==========================================

engine = create_engine(SQLALCHEMY_DATABASE_URI)
Session = sessionmaker(bind=engine)


# ==========================================
# Data models
# ==========================================

@dataclass
class LineItem:
    article_no: str
    quantity: int
    eta: str | None = None


@dataclass
class ParseResult:
    supplier: str
    customer_name: str
    order_date: str | None
    po_number: str | None
    items: list[LineItem] = field(default_factory=list)
    status: str = "ok"
    message: str = ""

    # Optional large banner shown on the review PDF
    review_banner_text: str | None = None
    review_banner_is_error: bool = False


@dataclass
class ReviewRow:
    excel_row: int
    article_no: str
    quantity: int
    eta: str | None


# ==========================================
# Small generic helpers
# ==========================================

def write_text_file(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def reset_work_folders() -> None:
    """
    Delete and recreate temporary work folders.
    """
    for folder in [INBOX_DIR, MAIL_ATTACHMENTS_DIR, REVIEW_DIR]:
        if folder.exists():
            shutil.rmtree(folder)
        folder.mkdir(parents=True, exist_ok=True)


def next_output_paths() -> tuple[Path, Path]:
    """
    Create unique output paths for one processing run.
    """
    token = uuid.uuid4().hex[:10]
    excel_path = OUTPUT_DIR / f"output_{token}.xlsx"
    summary_path = OUTPUT_DIR / f"summary_{token}.pdf"
    return excel_path, summary_path


def safe_filename(name: str | None, fallback: str) -> str:
    """
    Create a filesystem-safe filename.
    """
    if not name:
        return fallback

    cleaned = re.sub(r"[^A-Za-z0-9._ -]", "_", name).strip()
    return cleaned or fallback


# ==========================================
# PDF helpers
# ==========================================

def get_first_page_text(pdf_path: Path) -> str:
    try:
        with pdf_path.open("rb") as f:
            reader = PdfReader(f)
            if not reader.pages:
                return ""
            return reader.pages[0].extract_text() or ""
    except Exception:
        return ""


def get_full_pdf_text(pdf_path: Path) -> str:
    try:
        return extract_text(str(pdf_path)) or ""
    except Exception:
        return ""


def get_pdf_page_count(pdf_path: Path) -> int:
    """
    Return the page count of a PDF.
    Return 0 if the file could not be read.
    """
    try:
        with pdf_path.open("rb") as f:
            reader = PdfReader(f)
            return len(reader.pages)
    except Exception:
        return 0


# ==========================================
# Input collection
# Supports .pdf, .msg, .eml
# ==========================================

def move_uploaded_files_to_inbox() -> None:
    """
    Move all uploaded files into the incoming work folder.
    """
    for file in UPLOADS_DIR.iterdir():
        if file.is_file():
            shutil.move(str(file), str(INBOX_DIR / file.name))


def extract_pdfs_from_msg(msg_path: Path) -> list[Path]:
    """
    Extract PDF attachments from a .msg file.
    """
    if extract_msg is None:
        print("extract_msg is not installed, skipping .msg file")
        return []

    pdf_paths: list[Path] = []

    try:
        msg = extract_msg.Message(str(msg_path))
        target_dir = MAIL_ATTACHMENTS_DIR / msg_path.stem
        target_dir.mkdir(parents=True, exist_ok=True)

        for attachment in msg.attachments:
            try:
                attachment.save(customPath=str(target_dir))
            except Exception:
                continue

        for file in target_dir.iterdir():
            if file.is_file() and file.suffix.lower() == ".pdf":
                pdf_paths.append(file)

    except Exception as exc:
        print(f"Could not read .msg file {msg_path.name}: {exc}")

    return pdf_paths


def extract_pdfs_from_eml(eml_path: Path) -> list[Path]:
    """
    Extract PDF attachments from a standard .eml file.
    Supports multiple PDF attachments in one email.
    """
    pdf_paths: list[Path] = []

    try:
        with eml_path.open("rb") as f:
            message = BytesParser(policy=policy.default).parse(f)

        target_dir = MAIL_ATTACHMENTS_DIR / eml_path.stem
        target_dir.mkdir(parents=True, exist_ok=True)

        attachment_index = 0

        for part in message.iter_attachments():
            filename = part.get_filename()
            content_type = part.get_content_type()

            if filename:
                suffix = Path(filename).suffix.lower()
            else:
                guessed_ext = mimetypes.guess_extension(content_type or "") or ""
                suffix = guessed_ext.lower()

            is_pdf = suffix == ".pdf" or content_type == "application/pdf"
            if not is_pdf:
                continue

            attachment_index += 1

            safe_name = safe_filename(filename, f"attachment_{attachment_index}.pdf")
            if not safe_name.lower().endswith(".pdf"):
                safe_name += ".pdf"

            out_path = target_dir / safe_name

            payload = part.get_payload(decode=True)
            if payload is None:
                continue

            out_path.write_bytes(payload)
            pdf_paths.append(out_path)

    except Exception as exc:
        print(f"Could not read .eml file {eml_path.name}: {exc}")

    return pdf_paths


def collect_input_pdfs() -> list[Path]:
    """
    Collect PDFs from:
    - direct uploaded PDF files
    - Outlook .msg files
    - standard .eml email files

    One email can contain many PDF attachments.
    """
    reset_work_folders()
    move_uploaded_files_to_inbox()

    pdfs: list[Path] = []

    for file in INBOX_DIR.iterdir():
        if not file.is_file():
            continue

        suffix = file.suffix.lower()

        if suffix == ".pdf":
            pdfs.append(file)

        elif suffix == ".msg":
            pdfs.extend(extract_pdfs_from_msg(file))

        elif suffix == ".eml":
            pdfs.extend(extract_pdfs_from_eml(file))

    return pdfs


# ==========================================
# Supplier detection
# ==========================================

def detect_supplier(first_page_text: str) -> str:
    """
    Detect supplier from the first page text.
    """
    text = first_page_text.lower()

    if "sodaantarctica" in text:
        return "SodaAntarctica"

    if "bigcustomer" in text:
        return "BigCustomer"

    return "Unknown"


# ==========================================
# Supplier-specific parsers
# No shared date / PO / line parsing helpers.
# Each supplier owns its full parsing logic.
# ==========================================

def parse_soda_antarctica_pdf(pdf_path: Path) -> ParseResult:
    """
    Full SodaAntarctica parser.
    """
    full_text = get_full_pdf_text(pdf_path)
    text = full_text.replace("\r", "")
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    order_date = None
    po_number = None
    eta = None
    items: list[LineItem] = []

    # SodaAntarctica date logic
    for line in lines:
        match = re.search(r"\bDate:\s*([0-9]{1,4}[/-][0-9]{1,2}[/-][0-9]{1,4})", line, re.IGNORECASE)
        if match:
            order_date = match.group(1)
            break

    # SodaAntarctica PO logic
    for line in lines:
        match = re.search(r"\bOrder No:\s*([A-Za-z0-9-]+)", line, re.IGNORECASE)
        if match:
            po_number = match.group(1)
            break

    # SodaAntarctica ETA logic
    all_dates = re.findall(r"\b\d{2}-\d{2}-\d{4}\b", text)
    for date_value in all_dates:
        if date_value != order_date:
            eta = date_value
            break

    # SodaAntarctica line parsing logic
    article_lines: list[str] = []
    qty_values: list[int] = []

    for line in lines:
        if "A101" in line:
            article_lines.append(line)

        qty_match = re.search(r"\bEA:\s*(\d+)\b", line, re.IGNORECASE)
        if qty_match:
            qty_values.append(int(qty_match.group(1)))

    for i, article_line in enumerate(article_lines):
        article_match = re.search(r"\b([A-Z0-9-]{4,})\b", article_line)
        article_no = article_match.group(1) if article_match else article_line[:30]
        quantity = qty_values[i] if i < len(qty_values) else 0

        items.append(LineItem(article_no=article_no, quantity=quantity, eta=eta))

    if not items:
        items.append(LineItem(article_no="UNKNOWN-SA-ITEM", quantity=0, eta=eta))

    return ParseResult(
        supplier="SodaAntarctica",
        customer_name="SodaAntarctica&Co",
        order_date=order_date,
        po_number=po_number,
        items=items,
        status="ok",
        message="Parsed with SodaAntarctica parser",
    )


def parse_bigcustomer_pdf(pdf_path: Path) -> ParseResult:
    """
    Full BigCustomer parser.
    """
    full_text = get_full_pdf_text(pdf_path)
    text = full_text.replace("\r", "")
    lines = [line.strip() for line in text.splitlines() if line.strip()]

    order_date = None
    po_number = None
    eta = None
    items: list[LineItem] = []

    # BigCustomer date logic
    for line in lines:
        match = re.search(r"\bDate:\s*([0-9]{1,4}[/-][0-9]{1,2}[/-][0-9]{1,4})", line, re.IGNORECASE)
        if match:
            order_date = match.group(1)
            break

    # BigCustomer PO logic
    for line in lines:
        match = re.search(r"\bOrder No:\s*([A-Za-z0-9-]+)", line, re.IGNORECASE)
        if match:
            po_number = match.group(1)
            break

    # BigCustomer ETA logic
    all_dates = re.findall(r"\b\d{2}-\d{2}-\d{4}\b", text)
    for date_value in all_dates:
        if date_value != order_date:
            eta = date_value
            break

    # BigCustomer line parsing logic
    article_lines: list[str] = []
    qty_values: list[int] = []

    for line in lines:
        if "A101" in line:
            article_lines.append(line)

        qty_match = re.search(r"\bEA:\s*(\d+)\b", line, re.IGNORECASE)
        if qty_match:
            qty_values.append(int(qty_match.group(1)))

    for i, article_line in enumerate(article_lines):
        article_match = re.search(r"\b([A-Z0-9-]{4,})\b", article_line)
        article_no = article_match.group(1) if article_match else article_line[:30]
        quantity = qty_values[i] if i < len(qty_values) else 0

        items.append(LineItem(article_no=article_no, quantity=quantity, eta=eta))

    if not items:
        items.append(LineItem(article_no="UNKNOWN-BC-ITEM", quantity=0, eta=eta))

    return ParseResult(
        supplier="BigCustomer",
        customer_name="BigCustomer",
        order_date=order_date,
        po_number=po_number,
        items=items,
        status="ok",
        message="Parsed with BigCustomer parser",
    )


def parse_unknown_pdf(pdf_path: Path) -> ParseResult:
    """
    Fallback result for suppliers that are not activated.
    """
    return ParseResult(
        supplier="Unknown",
        customer_name="Unknown / Not activated",
        order_date=None,
        po_number=None,
        items=[],
        status="unknown",
        message="No matching parser found",
        review_banner_text="    unknown / not activated",
        review_banner_is_error=True,
    )


def parse_pdf(pdf_path: Path) -> ParseResult:
    """
    Detect supplier and run the correct parser.
    """
    first_page_text = get_first_page_text(pdf_path)
    supplier = detect_supplier(first_page_text)

    if supplier == "SodaAntarctica":
        return parse_soda_antarctica_pdf(pdf_path)

    if supplier == "BigCustomer":
        return parse_bigcustomer_pdf(pdf_path)

    return parse_unknown_pdf(pdf_path)


# ==========================================
# Excel output
# Shared because output format is common.
# ==========================================

def ensure_workbook(excel_path: Path) -> None:
    if excel_path.exists():
        return

    if DEFAULT_EXCEL_TEMPLATE.exists():
        shutil.copy(DEFAULT_EXCEL_TEMPLATE, excel_path)

        # Optional: style template sheet too
        wb = load_workbook(excel_path)
        ws = wb[DEFAULT_OUTPUT_SHEET]
        style_output_worksheet(ws)
        wb.save(excel_path)
        wb.close()
        return

    wb = Workbook()
    ws = wb.active
    assert ws is not None

    ws.title = DEFAULT_OUTPUT_SHEET
    ws["A1"] = "Order Date"
    ws["B1"] = "Customer"
    ws["C1"] = "PO Number"
    ws["D1"] = "Article No"
    ws["E1"] = "Qty"
    ws["F1"] = "ETA"

    style_output_worksheet(ws)

    wb.save(excel_path)
    wb.close()




def find_next_empty_row(ws) -> int:
    row = 2
    while ws[f"A{row}"].value is not None:
        row += 1
    return row


def append_parse_result_to_excel(excel_path: Path, result: ParseResult) -> list[ReviewRow]:
    """
    Write parsed data to Excel and return the used row numbers.
    """
    ensure_workbook(excel_path)

    wb = load_workbook(excel_path)
    ws = wb[DEFAULT_OUTPUT_SHEET]

    start_row = find_next_empty_row(ws)
    review_rows: list[ReviewRow] = []

    if not result.items:
        ws[f"A{start_row}"] = result.order_date
        ws[f"B{start_row}"] = result.customer_name
        ws[f"C{start_row}"] = result.po_number
        ws[f"D{start_row}"] = ""
        ws[f"E{start_row}"] = ""
        ws[f"F{start_row}"] = ""
    else:
        row = start_row

        for item in result.items:
            ws[f"A{row}"] = result.order_date
            ws[f"B{row}"] = result.customer_name
            ws[f"C{row}"] = result.po_number
            ws[f"D{row}"] = item.article_no
            ws[f"E{row}"] = item.quantity
            ws[f"F{row}"] = item.eta

            review_rows.append(
                ReviewRow(
                    excel_row=row,
                    article_no=item.article_no,
                    quantity=item.quantity,
                    eta=item.eta,
                )
            )

            row += 1

    wb.save(excel_path)
    wb.close()

    return review_rows


def style_output_worksheet(ws) -> None:
    """
    Apply simple styling to the output sheet.
    """
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Header row styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Column widths
    ws.column_dimensions["A"].width = 14   # Order Date
    ws.column_dimensions["B"].width = 24   # Customer
    ws.column_dimensions["C"].width = 18   # PO Number
    ws.column_dimensions["D"].width = 22   # Article No
    ws.column_dimensions["E"].width = 10   # Qty
    ws.column_dimensions["F"].width = 14   # ETA

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add filter to header row
    ws.auto_filter.ref = "A1:F1"

# ==========================================
# Review overlay helpers
# ==========================================

def draw_review_field(
    can,
    *,
    original_x: float,
    original_y: float,
    original_w: float,
    overlay_x: float,
    overlay_y: float,
    overlay_w: float,
    overlay_text: str,
    border_color,
    overlay_fill_color,
    height: float = 14,
    yellow_offset_x: float = 0,
    border_extra_top: float = 0,
    border_extra_bottom: float = 0,
) -> None:
    padding = 4

    frame_overlay_x = overlay_x
    yellow_x = overlay_x + yellow_offset_x

    if original_w <= 0:
        outer_x = frame_overlay_x - padding
        outer_y = overlay_y - 2 - border_extra_bottom
        outer_right = frame_overlay_x + overlay_w + padding
    else:
        outer_x = min(original_x, frame_overlay_x) - padding
        outer_y = min(original_y, overlay_y) - 2 - border_extra_bottom
        outer_right = max(original_x + original_w, frame_overlay_x + overlay_w) + padding

    outer_w = outer_right - outer_x
    outer_h = height + 4 + border_extra_top + border_extra_bottom

    can.setStrokeColor(border_color)
    can.setLineWidth(2)
    can.rect(outer_x, outer_y, outer_w, outer_h, fill=0, stroke=1)

    can.setFillColor(overlay_fill_color)
    can.rect(yellow_x, overlay_y, overlay_w, height, fill=1, stroke=0)

    can.setFillColorRGB(0, 0, 0)
    can.drawString(yellow_x + 3, overlay_y + 4, overlay_text)

def draw_diagonal_banner(
    can,
    text: str,
    *,
    fill_color,
    angle: float = 28,
    rect_x: float = 40,
    rect_y: float = 550,
    rect_w: float = 900,
    rect_h: float = 90,
    font_name: str = "Helvetica-Bold",
    font_size: int = 28,
    text_x: float = 90,
    text_y: float = 580,
) -> None:
    """
    Draw a long diagonal banner that fully crosses the page.

    The rectangle is intentionally oversized so the banner edges stay
    outside the visible page after rotation.
    """
    can.saveState()
    can.rotate(angle)

    can.setFillColor(fill_color)
    can.rect(rect_x, rect_y, rect_w, rect_h, fill=True, stroke=0)

    can.setFont(font_name, font_size)
    can.setFillColorRGB(0, 0, 0)
    can.drawString(text_x, text_y, text)

    can.restoreState()

def create_generic_status_overlay(
    banner_text: str,
    *,
    is_error: bool = True,
) -> io.BytesIO:
    """
    Create a simple overlay with only a large status banner.
    Used for unknown suppliers or invalid PDFs.
    """
    packet = io.BytesIO()
    can = canvas.Canvas(packet)

    draw_large_status_banner(can, banner_text, is_error=is_error)

    can.save()
    packet.seek(0)
    return packet

def draw_large_status_banner(
        can,
        text: str,
        *,
        is_error: bool = True,
    ) -> None:
    """
    Draw a large diagonal status banner for errors / unknown supplier cases.
    """
    if is_error:
        fill_color = Color(1, 0.85, 0.85, alpha=0.78)
    else:
        fill_color = Color(1, 1, 0.6, alpha=0.78)

    draw_diagonal_banner(
        can,
        text,
        fill_color=fill_color,
        angle=28,
        rect_x=40,
        rect_y=500,
        rect_w=950,
        rect_h=95,
        font_name="Helvetica-Bold",
        font_size=28,
        text_x=350,
        text_y=532,
    )

def draw_review_copy_banner(can) -> None:
    """
    Draw the standard long diagonal 'Review Copy' banner
    using the same style as the error banner.
    """
    fill_color = Color(1, 1, 0.6, alpha=0.78)

    draw_diagonal_banner(
        can,
        "Review Copy",
        fill_color=fill_color,
        angle=28,
        rect_x=40,
        rect_y=550,
        rect_w=950,   # same long length
        rect_h=95,    # same height as error banner
        font_name="Helvetica-Bold",
        font_size=28, # same font size as error banner
        text_x=350,
        text_y=580,
    )

# ==========================================
# Supplier-specific review overlays
# ==========================================

def create_soda_review_overlay(
    order_date: str | None,
    po_number: str | None,
    review_rows: list[ReviewRow],
    banner_text: str | None = None,
    banner_is_error: bool = False,
) -> io.BytesIO:
    packet = io.BytesIO()
    can = canvas.Canvas(packet)

    lightyellow70 = Color(1, 1, 0.6, alpha=0.7)
    border_color = Color(1, 0, 0, alpha=1)
    overlay_fill = Color(1, 1, 0.6, alpha=1)

    vertical_offset = -32

    for row in review_rows:
        y = 538 + vertical_offset

        draw_review_field(
            can,
            original_x=55,
            original_y=y,
            original_w=95,
            overlay_x=10,
            overlay_y=y,
            overlay_w=82,
            overlay_text=f"Art.nr: {row.article_no}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
            yellow_offset_x=60,
        )

        draw_review_field(
            can,
            original_x=0,
            original_y=0,
            original_w=0,
            overlay_x=210,
            overlay_y=y,
            overlay_w=70,
            overlay_text=f"Excelrow: {row.excel_row}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        draw_review_field(
            can,
            original_x=320,
            original_y=y,
            original_w=90,
            overlay_x=320,
            overlay_y=y,
            overlay_w=45,
            overlay_text=f"Qty: {row.quantity}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        draw_review_field(
            can,
            original_x=420,
            original_y=y,
            original_w=165,
            overlay_x=420,
            overlay_y=y,
            overlay_w=88,
            overlay_text=f"eta {row.eta}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        vertical_offset -= 18

    draw_review_field(
        can,
        original_x=40,
        original_y=610,
        original_w=155,
        overlay_x=40,
        overlay_y=610,
        overlay_w=155,
        overlay_text=f"Order nr: {po_number}",
        border_color=border_color,
        overlay_fill_color=overlay_fill,
        border_extra_bottom=25,
    )

    draw_review_field(
        can,
        original_x=215,
        original_y=217,
        original_w=135,
        overlay_x=215,
        overlay_y=217,
        overlay_w=135,
        overlay_text=f"Order date: {order_date}",
        border_color=border_color,
        overlay_fill_color=overlay_fill,
        border_extra_bottom=22,
    )

    draw_review_copy_banner(can)

    can.save()
    packet.seek(0)
    return packet


def create_bigcustomer_review_overlay(
    order_date: str | None,
    po_number: str | None,
    review_rows: list[ReviewRow],
    banner_text: str | None = None,
    banner_is_error: bool = False,
) -> io.BytesIO:
    packet = io.BytesIO()
    can = canvas.Canvas(packet)

    lightyellow70 = Color(1, 1, 0.6, alpha=0.7)
    border_color = Color(1, 0, 0, alpha=1)
    overlay_fill = Color(1, 1, 0.6, alpha=1)

    vertical_offset = -88

    for row in review_rows:
        y = 538 + vertical_offset

        draw_review_field(
            can,
            original_x=55,
            original_y=y,
            original_w=95,
            overlay_x=10,
            overlay_y=y,
            overlay_w=82,
            overlay_text=f"Art.nr: {row.article_no}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
            yellow_offset_x=60,
        )

        draw_review_field(
            can,
            original_x=0,
            original_y=0,
            original_w=0,
            overlay_x=210,
            overlay_y=y,
            overlay_w=70,
            overlay_text=f"Excelrow: {row.excel_row}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        draw_review_field(
            can,
            original_x=320,
            original_y=y,
            original_w=90,
            overlay_x=320,
            overlay_y=y,
            overlay_w=45,
            overlay_text=f"Qty: {row.quantity}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        draw_review_field(
            can,
            original_x=420,
            original_y=y,
            original_w=165,
            overlay_x=420,
            overlay_y=y,
            overlay_w=88,
            overlay_text=f"eta {row.eta}",
            border_color=border_color,
            overlay_fill_color=overlay_fill,
        )

        vertical_offset -= 18

    draw_review_field(
        can,
        original_x=360,
        original_y=700,
        original_w=185,
        overlay_x=360,
        overlay_y=700,
        overlay_w=185,
        overlay_text=f"Order number: {po_number}",
        border_color=border_color,
        overlay_fill_color=overlay_fill,
        border_extra_bottom=25,
    )

    draw_review_field(
        can,
        original_x=360,
        original_y=615,
        original_w=160,
        overlay_x=360,
        overlay_y=615,
        overlay_w=160,
        overlay_text=f"Order date: {order_date}",
        border_color=border_color,
        overlay_fill_color=overlay_fill,
        border_extra_top=22,
    )

    draw_review_copy_banner(can)

    can.save()
    packet.seek(0)
    return packet


# ==========================================
# Review PDF creation
# ==========================================

def create_review_pdf(
    original_pdf_path: Path,
    output_pdf_path: Path,
    result: ParseResult,
    review_rows: list[ReviewRow],
) -> Path:
    """
    Create a review PDF by overlaying the first page.
    """
    if result.supplier == "SodaAntarctica":
        overlay_pdf_bytes = create_soda_review_overlay(
            order_date=result.order_date,
            po_number=result.po_number,
            review_rows=review_rows,
            banner_text=result.review_banner_text,
            banner_is_error=result.review_banner_is_error,
        )

    elif result.supplier == "BigCustomer":
        overlay_pdf_bytes = create_bigcustomer_review_overlay(
            order_date=result.order_date,
            po_number=result.po_number,
            review_rows=review_rows,
            banner_text=result.review_banner_text,
            banner_is_error=result.review_banner_is_error,
        )

    else:
        banner_text = result.review_banner_text or result.message or "ERROR"
        overlay_pdf_bytes = create_generic_status_overlay(
            banner_text=banner_text,
            is_error=True,
        )

    overlay_reader = PdfReader(overlay_pdf_bytes)

    with original_pdf_path.open("rb") as original_file:
        original_reader = PdfReader(original_file)
        writer = PdfWriter()

        for i, page in enumerate(original_reader.pages):
            if i == 0 and overlay_reader.pages:
                page.merge_page(overlay_reader.pages[0])
            writer.add_page(page)

        with output_pdf_path.open("wb") as out_file:
            writer.write(out_file)

    return output_pdf_path


# ==========================================
# Summary PDF
# ==========================================

def create_summary_pdf(source_pdfs: list[Path], summary_output_path: Path) -> int:
    """
    Merge all review PDFs into one summary PDF.
    """
    writer = PdfWriter()
    count = 0

    for pdf_path in source_pdfs:
        writer.append(str(pdf_path))
        count += 1

    with summary_output_path.open("wb") as f:
        writer.write(f)

    return count


# ==========================================
# Main processing
# Single-page PDFs only
# ==========================================

def process_single_pdf(pdf_path: Path, excel_path: Path) -> Path:
    """
    Process one PDF.

    Rules:
    - only single-page PDFs are supported
    - if page count > 1, create error review PDF
    - if supplier is unknown, create review PDF with "unknown / not activated"
    - if parsing crashes, create review PDF with an error banner
    """
    page_count = get_pdf_page_count(pdf_path)

    if page_count == 0:
        result = ParseResult(
            supplier="Unknown",
            customer_name="Error",
            order_date=None,
            po_number=None,
            items=[],
            status="error",
            message="ERROR: Could not read PDF",
            review_banner_text="ERROR: Could not read PDF",
            review_banner_is_error=True,
        )
        review_rows: list[ReviewRow] = []

    elif page_count > 1:
        result = ParseResult(
            supplier="Unknown",
            customer_name="Error",
            order_date=None,
            po_number=None,
            items=[],
            status="error",
            message="ERROR: More than 1 page",
            review_banner_text="   ERROR: More than 1 page",
            review_banner_is_error=True,
        )
        review_rows = []

    else:
        try:
            result = parse_pdf(pdf_path)

            if result.status == "ok":
                review_rows = append_parse_result_to_excel(excel_path, result)
            else:
                review_rows = []

        except Exception as exc:
            short_error = str(exc)[:60]

            result = ParseResult(
                supplier="Unknown",
                customer_name="Error",
                order_date=None,
                po_number=None,
                items=[],
                status="error",
                message=f"ERROR: {short_error}",
                review_banner_text=f"ERROR: {short_error}",
                review_banner_is_error=True,
            )
            review_rows = []

    review_output_path = REVIEW_DIR / f"{pdf_path.stem}_review.pdf"

    create_review_pdf(
        original_pdf_path=pdf_path,
        output_pdf_path=review_output_path,
        result=result,
        review_rows=review_rows,
    )

    return review_output_path


def process_uploaded_files() -> tuple[int, Path, Path]:
    """
    Process all uploaded files.

    Supports:
    - PDF uploads
    - MSG email uploads with PDF attachments
    - EML email uploads with PDF attachments

    Only single-page PDFs are processed as valid orders.
    Multi-page PDFs still appear in summary.pdf with an error overlay.
    """
    input_pdfs = collect_input_pdfs()

    if not input_pdfs:
        raise RuntimeError("No input PDFs found in uploads, .msg attachments, or .eml attachments")

    excel_path, summary_path = next_output_paths()
    ensure_workbook(excel_path)

    review_pdfs: list[Path] = []

    for pdf in input_pdfs:
        review_pdf = process_single_pdf(pdf, excel_path)
        review_pdfs.append(review_pdf)
    


    merged_count = create_summary_pdf(review_pdfs, summary_path)

    write_text_file(EXCEL_PATH_TXT, excel_path.name)
    write_text_file(SUMMARY_PATH_TXT, summary_path.name)

    return merged_count, excel_path, summary_path


# ==========================================
# Job queue / worker
# ==========================================

def get_pending_job_slug() -> str | None:
    with Session.begin() as session:
        job = session.query(Job).filter_by(state="queued").first()

        if job is None:
            return None

        job.state = "processing"
        return job.slug


def mark_job_done(slug: str, result_value: int) -> None:
    with Session.begin() as session:
        session.query(Job).filter_by(slug=slug).update(
            {
                "state": "done",
                "result": result_value,
            }
        )


def mark_job_failed(slug: str) -> None:
    with Session.begin() as session:
        session.query(Job).filter_by(slug=slug).update(
            {
                "state": "failed",
                "result": -1,
            }
        )


def process_job(slug: str) -> None:
    print(f"Processing job: {slug}")

    try:
        merged_count, excel_path, summary_path = process_uploaded_files()

        mark_job_done(slug, merged_count)

        print(f"Done. Merged count = {merged_count}")
        print(f"Excel file:   {excel_path}")
        print(f"Summary file: {summary_path}")

    except Exception as exc:
        print(f"Job failed: {exc}")
        mark_job_failed(slug)


def run_worker_forever(poll_seconds: int = 1) -> None:
    print("Worker started")

    while True:
        slug = get_pending_job_slug()

        if slug:
            process_job(slug)
        else:
            time.sleep(poll_seconds)


# ==========================================
# Main entry point
# ==========================================

if __name__ == "__main__":
    run_worker_forever()
